''' 
Reads models specified in Excel into a Python object

@author Jonathan Karr, karr@mssm.edu
@date 3/22/2016
@author: Arthur Goldberg, Arthur.Goldberg@mssm.edu
'''

import warnings
with warnings.catch_warnings():
    warnings.simplefilter("ignore")
    from cobra import Metabolite as CobraMetabolite
    from cobra import Model as CobraModel
    from cobra import Reaction as CobraReaction
from openpyxl import load_workbook
import math
import re

from Sequential_WC_Simulator.multialgorithm.config import WC_SimulatorConfig
from Sequential_WC_Simulator.multialgorithm.model_representation import *
from Sequential_WC_Simulator.multialgorithm.submodels import submodel


#Reads model from Excel file into a Python object
def getModelFromExcel(filename, debug_option=False ):
    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", "Discarded range with reserved name", UserWarning)
        wb = load_workbook(filename = filename)

    #initialize model object
    model = Model( debug=debug_option )

    '''Read details from Excel'''
    #submodels
    ws = wb['Submodels']
    for iRow in range(2, ws.max_row + 1):
        # COMMENT(Arthur): make this more concise & data driven; column numbers very fragile
        # e.g., write ws.getNextRow() to get next row
        # and id = ws.get( 'ID' ), id = ws.get( 'Name' ), etc. where 'ID' refers to the column heading
        # even better, write  id, name, algorithm = ws.get( 'ID', 'Name', 'Algorithm' ) 
        id = ws.cell(row = iRow, column = 1).value
        name = ws.cell(row = iRow, column = 2).value
        algorithm = ws.cell(row = iRow, column = 3).value
        '''
        COMMENT(Arthur): 
        Must not instantiate submodels yet, as they may be objects running on other nodes, either for
        parallel execution of sequential simulations or parallel simulations.
        Rather, save them in submodel structures.

        if algorithm == 'FBA':
            subModel = FbaSubmodel( model, id = id, name = name)
        elif algorithm == 'SSA':
            subModel = SsaSubmodel( model, id = id, name = name)
        else:
            # COMMENT(Arthur): can use "".format()
            raise Exception('Undefined algorithm "%s" for submodel "%s"' % (algorithm, id))
        '''
        the_submodel = SubmodelSpecification( id, name, algorithm )
        model.submodels.append(the_submodel)
            
    #compartments
    ws = wb['Compartments']
    for iRow in range(2, ws.max_row + 1):
        model.compartments.append(Compartment(
            id = ws.cell(row = iRow, column = 1).value,
            name = ws.cell(row = iRow, column = 2).value,
            initialVolume = float(ws.cell(row = iRow, column = 3).value),
            comments = ws.cell(row = iRow, column = 4).value,
            ))
    
    #species
    ws = wb['Species']
    for iRow in range(2, ws.max_row + 1):
        mwStr = ws.cell(row = iRow, column = 5).value
        if mwStr:
            mw = float(mwStr)
        else:
            mw = None
            
        chargeStr = ws.cell(row = iRow, column = 6).value
        if chargeStr:
            charge = float(chargeStr)
        else:
            charge = None
    
        model.species.append(Species(
            id = ws.cell(row = iRow, column = 1).value,
            name = ws.cell(row = iRow, column = 2).value,
            structure = ws.cell(row = iRow, column = 3).value,
            empiricalFormula = ws.cell(row = iRow, column = 4).value,
            molecularWeight = mw,
            charge = charge,
            type = ws.cell(row = iRow, column = 7).value,
            concentrations = [
                Concentration(compartment = 'c', value = float(ws.cell(row = iRow, column = 8).value or 0)),
                Concentration(compartment = 'e', value = float(ws.cell(row = iRow, column = 9).value or 0)),
                ],
            crossRefs = [
                CrossReference(
                    source = ws.cell(row = iRow, column = 10).value, 
                    id = ws.cell(row = iRow, column = 11).value,
                    ),
                ],
            comments = ws.cell(row = iRow, column = 12).value,
            ))
            
    #reactions
    ws = wb['Reactions']
    for iRow in range(2, ws.max_row + 1):
        stoichiometry = parseStoichiometry(ws.cell(row = iRow, column = 4).value)
        
        rateLawStr = ws.cell(row = iRow, column = 6).value
        if rateLawStr:
            rateLaw = RateLaw(rateLawStr)
        else:
            rateLaw = None
        
        model.reactions.append(Reaction(
            id = ws.cell(row = iRow, column = 1).value,
            name = ws.cell(row = iRow, column = 2).value,
            submodel_spec = ws.cell(row = iRow, column = 3).value,
            reversible = stoichiometry['reversible'],
            participants = stoichiometry['participants'],
            enzyme = ws.cell(row = iRow, column = 5).value,
            rateLaw = rateLaw,
            vmax = ws.cell(row = iRow, column = 7).value,
            km = ws.cell(row = iRow, column = 8).value,
            crossRefs = [
                CrossReference(
                    source = ws.cell(row = iRow, column = 9).value, 
                    id = ws.cell(row = iRow, column = 10).value,
                    ),
                ],
            comments = ws.cell(row = iRow, column = 11).value,
            ))
            
    #parameters
    ws = wb['Parameters']
    for iRow in range(2, ws.max_row + 1):
        model.parameters.append(Parameter(
            id = ws.cell(row = iRow, column = 1).value,
            name = ws.cell(row = iRow, column = 2).value,
            submodel_spec = ws.cell(row = iRow, column = 3).value,
            value = float(ws.cell(row = iRow, column = 4).value),
            units = ws.cell(row = iRow, column = 5).value,
            comments = ws.cell(row = iRow, column = 6).value,
            ))
            
    #references
    ws = wb['References']
    for iRow in range(2, ws.max_row + 1):
        model.references.append(Reference(
            id = ws.cell(row = iRow, column = 1).value,
            name = ws.cell(row = iRow, column = 2).value,
            crossRefs = [
                CrossReference(
                    source = ws.cell(row = iRow, column = 3).value, 
                    id = ws.cell(row = iRow, column = 4).value,
                    ),
                ],
            comments = ws.cell(row = iRow, column = 5).value,
            ))
            
    '''set component indices'''
    model.setComponentIndices()
            
    '''deserialize references'''
    # ensure that all components are defined
    undefinedComponents = []
    
    #species concentration 
    for species in model.species:
        for conc in species.concentrations:
            id = conc.compartment            
            obj = model.getComponentById(id, model.compartments)
            if id and obj is None:
                undefinedComponents.append(id)
            conc.compartment = obj                
            
    #reaction submodel, participant species, participant compartments, enzymes
    for reaction in model.reactions:
        id = reaction.submodel_spec
        obj = model.getComponentById(id, model.submodels)
        if id and obj is None:
            undefinedComponents.append(id)
        reaction.submodel_spec = obj
        
        for part in reaction.participants:
            id = part.species            
            obj = model.getComponentById(id, model.species)
            if id and obj is None:               
                undefinedComponents.append(id)
            part.species = obj
                
            id = part.compartment            
            obj = model.getComponentById(id, model.compartments)
            if id and obj is None:
                undefinedComponents.append(id)
            part.compartment = obj
            
            part.calcIdName()
        
        id = reaction.enzyme
        obj = model.getComponentById(id, model.species)
        if id and obj is None:
            undefinedComponents.append(id)
        reaction.enzyme = obj

    #parameter submodels
    for param in model.parameters:
        id = param.submodel_spec
        if id:
            obj = model.getComponentById(id, model.submodels)
            if obj is None:
                undefinedComponents.append(id)
            param.submodel_spec = obj

    if len(undefinedComponents) > 0:
        undefinedComponents = list(set(undefinedComponents))
        undefinedComponents.sort()
        raise Exception('Undefined components:\n- %s' % ('\n- '.join(undefinedComponents)))
        
    ''' Assemble back references'''
    for submodel_spec in model.submodels:
        submodel_spec.reactions = []
        submodel_spec.species = []
        submodel_spec.parameters = []
    for rxn in model.reactions:
        rxn.submodel_spec.reactions.append(rxn)
        for part in rxn.participants:
            rxn.submodel_spec.species.append('%s[%s]' % (part.species.id, part.compartment.id))
        if rxn.enzyme:
            rxn.submodel_spec.species.append('%s[%s]' % (rxn.enzyme.id, 'c'))
        if rxn.rateLaw:
            rxn.submodel_spec.species += rxn.rateLaw.getModifiers(model.species, model.compartments)
    
    for param in model.parameters:
        if param.submodel_spec:
            param.submodel_spec.parameters.append(param)
            
    for subModel in model.submodels:
        speciesStrArr = list(set(subModel.species))
        speciesStrArr.sort()
        subModel.species = []
        for index, speciesStr in enumerate(speciesStrArr):
            speciesId, compId = speciesStr.split('[')
            compId = compId[0:-1]
            speciesComp = SpeciesCompartment(
                index = index,
                species = model.getComponentById(speciesId, model.species),
                compartment = model.getComponentById(compId, model.compartments),
                )
            speciesComp.calcIdName()
            subModel.species.append(speciesComp)
            
    '''Transcode rate laws'''
    for rxn in model.reactions:
        if rxn.rateLaw:
            rxn.rateLaw.transcode(model.species, model.compartments)
        
    return model
    
#Parse a string representing the stoichiometry of a reaction into a Python object
def parseStoichiometry(rxnStr):
    #Split stoichiometry in to global compartment, left-hand side, right-hand side, reversibility indictor
    '''
    COMMENT(Arthur): while RE parsers can be written quickly, for long-term use I think they're generally a bad idea 
    because they're difficult to read and quite brittle. That is, changes to the language can be quite difficult
    to implement in the parser. Also, RE parsers tend to be buggie, i.e., the patterns accept or fail to accept 
    unanticipated edge cases. 
    Instead, I recommend that we get in the habit of building grammer-based parsers. In exchange for taking
    longer to built (but less longer as one gets experienced), we'll get parsers that are much easier to 
    understand, test, and change. Also, they generate much better errors and the languages they accept will 
    be easier to explain and document.
    Several python based parsers are described at https://wiki.python.org/moin/LanguageParsing.
    '''
    rxnMatch = re.match('(?P<compartment>\[([a-z])\]: )?(?P<lhs>((\(\d*\.?\d*([e][-+]?[0-9]+)?\) )?[a-z0-9\-_]+(\[[a-z]\])? \+ )*(\(\d*\.?\d*([e][-+]?[0-9]+)?\) )?[a-z0-9\-_]+(\[[a-z]\])?) (?P<direction>[<]?)==> (?P<rhs>((\(\d*\.?\d*([e][-+]?[0-9]+)?\) )?[a-z0-9\-_]+(\[[a-z]\])? \+ )*(\(\d*\.?\d*([e][-+]?[0-9]+)?\) )?[a-z0-9\-_]+(\[[a-z]\])?)', rxnStr, flags=re.I)
    if rxnMatch is None:
        raise Exception('Invalid stoichiometry: %s' % rxnStr)
        
    #Determine reversiblity
    rxnDict = rxnMatch.groupdict()
    reversible = rxnDict['direction'] == '<'
    
    #Determine if global compartment for reaction was specified
    if rxnDict['compartment'] is None:
        globalComp = None
    else:
        globalComp = re.match('\[(?P<compartment>[a-z])\]', rxnDict['compartment'], flags=re.I).groupdict()['compartment']
    
    #initialize array of reaction participants
    participants = []
    
    #Parse left-hand side
    for rxnPartStr in rxnDict['lhs'].split(' + '):
        rxnPartDict = re.match('(\((?P<coefficient>\d*\.?\d*([e][-+]?[0-9]+)?)\) )?(?P<species>[a-z0-9\-_]+)(\[(?P<compartment>[a-z])\])?', rxnPartStr, flags=re.I).groupdict()
                
        species = rxnPartDict['species']
        compartment = rxnPartDict['compartment'] or globalComp
        coefficient = float(rxnPartDict['coefficient'] or 1)
    
        participants.append(ReactionParticipant(
            species = species,
            compartment = compartment,
            coefficient = -coefficient,
            ))
            
    #Parse right-hand side
    for rxnPartStr in rxnDict['rhs'].split(' + '):
        '''COMMENT(Arthur): same RE as above; reuse'''
        rxnPartDict = re.match('(\((?P<coefficient>\d*\.?\d*([e][-+]?[0-9]+)?)\) )?(?P<species>[a-z0-9\-_]+)(\[(?P<compartment>[a-z])\])?', rxnPartStr, flags=re.I).groupdict()
        
        species = rxnPartDict['species']
        compartment = rxnPartDict['compartment'] or globalComp
        coefficient = float(rxnPartDict['coefficient'] or 1)
        
        participants.append(ReactionParticipant(
            species = species,
            compartment = compartment,
            coefficient = coefficient,
            ))

    return {
        'reversible': reversible,
        'participants': participants,
        }
